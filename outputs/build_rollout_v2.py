"""
Builds the complete API Hunting ABM Rollout Schedule v2.

Includes:
  - All existing campaigns from v1 (P0 + 2 P1 + 12 P2 use cases)
  - 7 previously missing use cases identified from the Dream List
  - Sales POC owners and launch dates following the 2/week cadence
  - A change-log sheet for POC visibility
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/naitikchavda/Event Auto push/smartlead-kit/outputs/Campaign_Rollout_Schedule_v2_Complete.xlsx"

# (week, launch_day, campaign, priority, companies, poc, status, source, notes)
# status: Live | In Progress | Planned | New
ROWS = [
    # P0
    ("Week 1", "Mon Apr 14", "Passive Pipeline (Full blast)", "P0", 946,
     "All SDRs", "Live", "v1", ""),
    # P1
    ("Week 1", "Thu Apr 17", "Active Pipeline — Customer Loyalty", "P1", "TBD",
     "Akshat", "Live", "v1", "Stalled deals, Gaurav's list"),
    ("Week 2", "Mon Apr 21", "Active Pipeline — Fintech/BFSI", "P1", "TBD",
     "Gaurav/Avipsa", "Live", "v1", "Stalled deals"),
    # P2 existing 12
    ("Week 2", "Thu Apr 24", "Health & Wellness Apps (B2C)", "P2", 274,
     "Kishan", "In Progress", "v1", ""),
    ("Week 3", "Mon Apr 28", "Digital / Neo Banking", "P2", 277,
     "Gaurav/Avipsa", "In Progress", "v1", ""),
    ("Week 3", "Thu May 1", "E-Wallets & Payment Apps", "P2", 84,
     "Gaurav/Avipsa", "Planned", "v1", ""),
    ("Week 4", "Mon May 5", "Employee R&R Platforms", "P2", 123,
     "Kishan", "Planned", "v1", ""),
    ("Week 4", "Thu May 8", "Market Research Agencies", "P2", 602,
     "Rakesh", "Planned", "v1", ""),
    ("Week 5", "Mon May 12", "Employee Fringe Benefits", "P2", 29,
     "Kishan", "Planned", "v1", ""),
    ("Week 5", "Thu May 15", "Customer Loyalty Platforms", "P2", 99,
     "Akshat", "Planned", "v1", ""),
    ("Week 6", "Mon May 19", "Channel Loyalty Platforms", "P2", 70,
     "Akshat", "Planned", "v1", ""),
    ("Week 6", "Thu May 22", "Task-Based Rewards / Play-to-Earn", "P2", 66,
     "Akshat", "Planned", "v1", ""),
    ("Week 7", "Mon May 26", "Cashback Platforms", "P2", 75,
     "Rakesh", "Planned", "v1", ""),
    ("Week 7", "Thu May 29", "Remittance + Fintech & BFSI Global", "P2", 170,
     "Gaurav/Avipsa", "Planned", "v1", "Combined: Remittance 35 + Fintech BFSI Global 30 + wider pull"),
    # P2 NEW additions (previously missing)
    ("Week 8", "Mon Jun 2", "Customer Referral Platforms", "P2", 42,
     "Akshat", "New", "v2", "Added from Dream List gap"),
    ("Week 8", "Thu Jun 5", "Employee Referral Platforms", "P2", 43,
     "Kishan", "New", "v2", "Added from Dream List gap"),
    ("Week 9", "Mon Jun 9", "Customer Advocacy Platforms", "P2", 22,
     "Akshat", "New", "v2", "Added from Dream List gap"),
    ("Week 9", "Thu Jun 12", "Employee Advocacy Platforms", "P2", 11,
     "Kishan", "New", "v2", "Added from Dream List gap"),
    ("Week 10", "Mon Jun 16", "Fintech & BFSI India", "P2", 20,
     "Shailey/Gaurav", "New", "v2", "Added from Dream List gap; distinct from Global"),
    ("Week 10", "Thu Jun 19", "Survey Platforms + Online Gaming/Casino Loyalty", "P2", 21,
     "Rakesh + Gaurav", "New", "v2",
     "Combined: Survey 9 + Online Gaming/Casino cluster (~12, was 11 fragmented labels in Dream List)"),
]

CHANGELOG = [
    ("Date", "Change", "Owner"),
    ("2026-04-29", "v2 created — 6 launch slots added covering 7 missing use cases", "Naitik"),
    ("2026-04-29", "Customer Referral Platforms — 42 cos — Akshat — Mon Jun 2", "Naitik"),
    ("2026-04-29", "Employee Referral Platforms — 43 cos — Kishan — Thu Jun 5", "Naitik"),
    ("2026-04-29", "Customer Advocacy Platforms — 22 cos — Akshat — Mon Jun 9", "Naitik"),
    ("2026-04-29", "Employee Advocacy Platforms — 11 cos — Kishan — Thu Jun 12", "Naitik"),
    ("2026-04-29", "Fintech & BFSI India — 20 cos — Shailey/Gaurav — Mon Jun 16", "Naitik"),
    ("2026-04-29", "Survey + Online Gaming/Casino Loyalty — 21 cos — Rakesh + Gaurav — Thu Jun 19", "Naitik"),
    ("2026-04-29", "Open: 247 unassigned rows in Dream List need triage before launch", "Naitik"),
]

POC_NOTES = [
    ("Sales POC", "Use Cases Owned", "Total Companies", "Launch Window"),
    ("Akshat", "Customer Loyalty, Channel Loyalty, Task-Based Rewards, Customer Referral, Customer Advocacy",
     99 + 70 + 66 + 42 + 22, "May 15 to Jun 9"),
    ("Kishan", "Health & Wellness, Employee R&R, Employee Fringe Benefits, Employee Referral, Employee Advocacy",
     274 + 123 + 29 + 43 + 11, "Apr 24 to Jun 12"),
    ("Gaurav/Avipsa", "Digital Neo Banking, E-Wallets, Remittance + Fintech BFSI Global, Online Gaming/Casino (combo)",
     277 + 84 + 170, "Apr 28 to Jun 19"),
    ("Shailey/Gaurav", "Fintech & BFSI India", 20, "Jun 16"),
    ("Rakesh", "Market Research, Cashback, Survey Platforms (combo)", 602 + 75, "May 8 to Jun 19"),
]


def main():
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Rollout Schedule
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Rollout Schedule"

    title = "API HUNTING ABM — CAMPAIGN ROLLOUT v2 (COMPLETE)"
    sub = "2 sequences live per week (Mon + Thu). All 22 Dream List use cases covered."

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = sub
    ws["A2"].font = Font(italic=True, color="555555")
    ws.merge_cells("A2:I2")

    headers = ["Week", "Launch Day", "Campaign", "Priority",
               "# Companies", "Sales POC", "Status", "Source", "Notes"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_fill = {
        "Live": PatternFill("solid", fgColor="C6EFCE"),
        "In Progress": PatternFill("solid", fgColor="FFEB9C"),
        "Planned": PatternFill("solid", fgColor="DDEBF7"),
        "New": PatternFill("solid", fgColor="F8CBAD"),
    }
    priority_fill = {
        "P0": PatternFill("solid", fgColor="C00000"),
        "P1": PatternFill("solid", fgColor="ED7D31"),
        "P2": PatternFill("solid", fgColor="70AD47"),
    }

    for ridx, row in enumerate(ROWS, start=5):
        for cidx, val in enumerate(row, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        # priority colour
        pcell = ws.cell(row=ridx, column=4)
        pcell.fill = priority_fill.get(row[3], PatternFill())
        pcell.font = Font(bold=True, color="FFFFFF")
        pcell.alignment = Alignment(horizontal="center", vertical="center")
        # status colour (whole row tinted by status, but only the status column bolded)
        scell = ws.cell(row=ridx, column=7)
        scell.fill = status_fill.get(row[6], PatternFill())
        scell.font = Font(bold=True)
        scell.alignment = Alignment(horizontal="center", vertical="center")
        # New rows: also tint columns A-C lightly so they stand out
        if row[6] == "New":
            for c in range(1, 4):
                ws.cell(row=ridx, column=c).fill = PatternFill("solid", fgColor="FFF2CC")

    # column widths
    widths = [10, 14, 50, 10, 13, 22, 14, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # Total row
    total_row = len(ROWS) + 5
    ws.cell(row=total_row, column=3, value="TOTAL (P0 + P1 + P2)").font = Font(bold=True)
    numeric = sum(r[4] for r in ROWS if isinstance(r[4], int))
    ws.cell(row=total_row, column=5, value=numeric).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value="P1 contact counts pending from Gaurav").font = Font(italic=True)
    for c in range(1, 10):
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")

    # ------------------------------------------------------------------
    # Sheet 2: POC Summary
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("POC Summary")
    ws2["A1"] = "OWNERSHIP BY SALES POC"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 22

    for ridx, row in enumerate(POC_NOTES, start=3):
        for cidx, val in enumerate(row, start=1):
            cell = ws2.cell(row=ridx, column=cidx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if ridx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, w in enumerate([18, 70, 18, 28], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    # ------------------------------------------------------------------
    # Sheet 3: Change Log
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Change Log")
    ws3["A1"] = "CHANGE LOG — v1 → v2"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws3.merge_cells("A1:C1")
    ws3.row_dimensions[1].height = 22

    for ridx, row in enumerate(CHANGELOG, start=3):
        for cidx, val in enumerate(row, start=1):
            cell = ws3.cell(row=ridx, column=cidx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if ridx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, w in enumerate([14, 90, 14], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A4"

    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Rows: {len(ROWS)}  Total companies (P0+P2): {numeric}")


if __name__ == "__main__":
    main()
