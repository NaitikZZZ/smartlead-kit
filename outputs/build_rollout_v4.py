"""
v4 of the API Hunting ABM Rollout Schedule.

Changes from v3:
  - Coverage Map now lists ALL 30 Dream List use case labels exactly as they
    appear in the source (typos preserved), so POCs can verify nothing was
    silently dropped or renamed.
  - Online Gaming & Casino Loyalty gets its own dedicated launch slot
    (10 sub-labels + Gaming gift card re-sellers = 11 companies).
  - Survey Platforms gets its own dedicated slot (no longer combined).
  - BAAS (1 co, ITIO) explicitly listed as folded into Remittance + Fintech BFSI Global.
  - 3 currently LIVE campaigns marked: Passive Deal, Customer Loyalty (Global API),
    Cashback Platforms.
  - ABM Touchpoints sheet retained with Smartlead numbers.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/naitikchavda/Event Auto push/smartlead-kit/outputs/Campaign_Rollout_Schedule_v4_Complete.xlsx"

# (week, launch_day, campaign, priority, companies, poc, status, source, notes)
ROWS = [
    # P0
    ("Week 1", "Mon Apr 14", "Global API — Passive Deal (Full blast)", "P0", 946,
     "All SDRs", "Live", "v1",
     "Smartlead: P0-C2-Passive-New-Deals-Apr2026 (id 3178228) — ACTIVE"),
    # P1
    ("Week 1", "Thu Apr 17", "Customer Loyalty — Global API", "P1", 278,
     "Gaurav/Akshat", "Live", "v1",
     "Smartlead: P0_API_CUSTOMLOYALTY_GLOBA (id 3222599) — ACTIVE. "
     "Stalled deals + dream list combined."),
    ("Week 2", "Mon Apr 21", "Active Pipeline — Fintech/BFSI", "P1", "TBD",
     "Gaurav/Avipsa", "Planned", "v1", "Stalled deals; counts pending"),
    # P2 — covering Dream List use cases in the order they appear
    ("Week 2", "Thu Apr 24", "Health & Wellness Apps (B2C)", "P2", 267,
     "Kishan", "In Progress", "v1",
     "Dream List label: 'Health & Welness Apps (B2C)' (typo preserved at source)"),
    ("Week 3", "Mon Apr 28", "Digital / Neo Banking", "P2", 134,
     "Gaurav/Avipsa", "In Progress", "v1", ""),
    ("Week 3", "Thu May 1", "E-Wallets & Payment Apps", "P2", 50,
     "Gaurav/Avipsa", "Planned", "v1", "Dream List label: 'E-wallets/Payment Apps'"),
    ("Week 4", "Mon May 5", "Employee R&R Platforms", "P2", 121,
     "Kishan", "Planned", "v1",
     "Dream List label: 'Employee Rewards and Engagement platforms'"),
    ("Week 4", "Thu May 8", "Market Research Agencies", "P2", 602,
     "Rakesh", "Planned", "v1", ""),
    ("Week 5", "Mon May 12", "Employee Fringe Benefits Platforms", "P2", 29,
     "Kishan", "Planned", "v1", ""),
    ("Week 5", "Thu May 15", "Customer Loyalty Platforms (Dream)", "P2", 98,
     "Akshat", "Planned", "v1",
     "Distinct from the Live P1 Customer Loyalty (Global API) running today"),
    ("Week 6", "Mon May 19", "Channel Loyalty Platforms", "P2", 70,
     "Akshat", "Planned", "v1", ""),
    ("Week 6", "Thu May 22", "Task-Based Rewards / Play-to-Earn", "P2", 59,
     "Akshat", "Planned", "v1", "Dream List label: 'Task Based Rewards/Play to Earn'"),
    # Cashback now LIVE
    ("Week 7", "Mon May 26", "Cashback Platforms", "P2", 71,
     "Rakesh + Gaurav", "Live", "v1",
     "Smartlead: P0_API_CASHBACK_GLOBA (id 3222821) — ACTIVE since Apr 23. "
     "Launched ahead of schedule. Dream List label: 'Cashback plaforms'"),
    ("Week 7", "Thu May 29", "Remittance + Fintech & BFSI Global (+BAAS)", "P2", 66,
     "Gaurav/Avipsa", "Planned", "v1",
     "Combined: Remittance (35) + Fintech & BFSI Global (30) + BAAS (1, ITIO)"),
    # P2 NEW additions (from Dream List gap)
    ("Week 8", "Mon Jun 2", "Customer Referral Platforms", "P2", 42,
     "Akshat", "New", "v2", "Added from Dream List gap"),
    ("Week 8", "Thu Jun 5", "Employee Referral Platforms", "P2", 43,
     "Kishan", "New", "v2", "Added from Dream List gap"),
    ("Week 9", "Mon Jun 9", "Customer Advocacy Platforms", "P2", 22,
     "Akshat", "New", "v2", "Added from Dream List gap"),
    ("Week 9", "Thu Jun 12", "Employee Advocacy Platforms", "P2", 11,
     "Kishan", "New", "v2",
     "Added from Dream List gap. Source label: 'Employe Advocacy platforms'"),
    ("Week 10", "Mon Jun 16", "Fintech & BFSI India", "P2", 20,
     "Shailey/Gaurav", "New", "v2",
     "Added from Dream List gap; distinct from Global"),
    ("Week 10", "Thu Jun 19", "Online Gaming & Casino Loyalty", "P2", 11,
     "Gaurav", "New", "v2",
     "Combined slot covering 10 fragmented Dream List labels (1 co each) + "
     "Gaming gift card re-sellers (1, Gameship). Sub-labels: Casino & Resort Loyalty Program; "
     "Casino & Resort Operator (Loyalty Program); Online & Land-based Casino Loyalty Program; "
     "Online Casino; Online Casino & Sportsbook; Online iGaming Platform; "
     "Online Poker & Casino Platform; Online Sportsbook & Casino; "
     "Online Sweepstakes Casino; Social / Online Casino; Gaming gift card re-sellers"),
    ("Week 11", "Mon Jun 23", "Survey Platforms", "P2", 9,
     "Rakesh", "New", "v2",
     "Added from Dream List gap. Source label: 'Survery Platforms'"),
]

# Coverage Map — every Dream List label preserved as-is from source
# (label, cos, mapped_to_slot, launch, poc, status)
COVERAGE = [
    ("Dream List Label (as-is in source)", "Cos", "Mapped to Rollout Slot",
     "Launch", "POC", "Status"),
    ("BAAS", 1, "Remittance + Fintech & BFSI Global (+BAAS)",
     "Thu May 29", "Gaurav/Avipsa", "Folded in"),
    ("Cashback plaforms", 71, "Cashback Platforms",
     "Mon May 26", "Rakesh + Gaurav", "LIVE since Apr 23"),
    ("Casino & Resort Loyalty Program", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Casino & Resort Operator (Loyalty Program)", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Channel Loyalty Platforms", 70, "Channel Loyalty Platforms",
     "Mon May 19", "Akshat", "Planned"),
    ("Customer Advocacy Platforms", 22, "Customer Advocacy Platforms",
     "Mon Jun 9", "Akshat", "NEW slot"),
    ("Customer Loyalty Platforms", 98, "Customer Loyalty Platforms (Dream)",
     "Thu May 15", "Akshat", "Planned (Live as P1 today, separate scope)"),
    ("Customer Referral Platforms", 42, "Customer Referral Platforms",
     "Mon Jun 2", "Akshat", "NEW slot"),
    ("Digital / Neo Banking", 134, "Digital / Neo Banking",
     "Mon Apr 28", "Gaurav/Avipsa", "In Progress"),
    ("Employe Advocacy platforms", 11, "Employee Advocacy Platforms",
     "Thu Jun 12", "Kishan", "NEW slot"),
    ("Employee Fringe Benefits Platforms", 29, "Employee Fringe Benefits Platforms",
     "Mon May 12", "Kishan", "Planned"),
    ("Employee Referral Platforms", 43, "Employee Referral Platforms",
     "Thu Jun 5", "Kishan", "NEW slot"),
    ("Employee Rewards and Engagement platforms", 121, "Employee R&R Platforms",
     "Mon May 5", "Kishan", "Planned"),
    ("E-wallets/Payment Apps", 50, "E-Wallets & Payment Apps",
     "Thu May 1", "Gaurav/Avipsa", "Planned"),
    ("Fintech & BFSI Global", 30, "Remittance + Fintech & BFSI Global (+BAAS)",
     "Thu May 29", "Gaurav/Avipsa", "Planned"),
    ("Fintech & BFSI India", 20, "Fintech & BFSI India",
     "Mon Jun 16", "Shailey/Gaurav", "NEW slot"),
    ("Gaming gift card re-sellers", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "Folded in"),
    ("Health & Welness Apps (B2C)", 267, "Health & Wellness Apps (B2C)",
     "Thu Apr 24", "Kishan", "In Progress"),
    ("Market Research Agencies", 602, "Market Research Agencies",
     "Thu May 8", "Rakesh", "Planned"),
    ("Online & Land-based Casino Loyalty Program", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Online Casino", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Online Casino & Sportsbook", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Online iGaming Platform", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Online Poker & Casino Platform", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Online Sportsbook & Casino", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Online Sweepstakes Casino", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Remittance", 35, "Remittance + Fintech & BFSI Global (+BAAS)",
     "Thu May 29", "Gaurav/Avipsa", "Planned"),
    ("Social / Online Casino", 1, "Online Gaming & Casino Loyalty",
     "Thu Jun 19", "Gaurav", "NEW slot — combined"),
    ("Survery Platforms", 9, "Survey Platforms",
     "Mon Jun 23", "Rakesh", "NEW slot"),
    ("Task Based Rewards/Play to Earn", 59, "Task-Based Rewards / Play-to-Earn",
     "Thu May 22", "Akshat", "Planned"),
    ("(Unassigned in Dream List — 247 rows)", 247, "TRIAGE NEEDED",
     "Pre-launch", "Naitik", "Open: needs sorting before any of the above slots"),
]

# ABM Touchpoints — only for Live campaigns
TOUCHPOINTS = [
    ("Campaign", "Smartlead Campaign ID", "Status",
     "Leads in SL", "Emails Sent", "Replies", "Reply %", "Bounces", "Unsubs",
     "LI Touches\n(HeyReach)", "Calls\n(HubSpot)", "Total\nTouchpoints",
     "Notes"),
    ("Global API — Passive Deal", "3178228", "ACTIVE since Apr 14",
     "264 / ~633 (sample)", "500 / 1265 (sample)", 6, "1.2%", 1, 0,
     "TBD — pull from HeyReach", "TBD — pull from HubSpot", "TBD",
     "Smartlead figures from first stats page (500 of 1265 records). "
     "Open/click are 0 because tracking is intentionally disabled. "
     "Replies and bounces are real."),
    ("Customer Loyalty — Global API", "3222599", "ACTIVE since Apr 23",
     142, 278, 4, "1.4%", 0, 0,
     "TBD — pull from HeyReach", "TBD — pull from HubSpot", "TBD",
     "Full data (278 of 278). Tracking disabled, so open/click are 0."),
    ("Cashback Platforms — Global API", "3222821", "ACTIVE since Apr 23",
     39, 77, 1, "1.3%", 0, 0,
     "TBD — pull from HeyReach", "TBD — pull from HubSpot", "TBD",
     "Full data (77 of 77). Tracking disabled, so open/click are 0."),
]

POC_NOTES = [
    ("Sales POC", "Use Cases Owned", "Total Companies", "Launch Window"),
    ("Akshat", "Customer Loyalty (Dream), Channel Loyalty, Task-Based Rewards, "
     "Customer Referral, Customer Advocacy",
     98 + 70 + 59 + 42 + 22, "May 15 to Jun 9"),
    ("Kishan", "Health & Wellness, Employee R&R, Employee Fringe Benefits, "
     "Employee Referral, Employee Advocacy",
     267 + 121 + 29 + 43 + 11, "Apr 24 to Jun 12"),
    ("Gaurav/Avipsa", "Digital Neo Banking, E-Wallets, Remittance + Fintech BFSI Global "
     "(+BAAS), Online Gaming/Casino, Cashback (co-owned), Customer Loyalty Live (co-owned)",
     134 + 50 + 66 + 11, "Apr 23 to Jun 19"),
    ("Shailey/Gaurav", "Fintech & BFSI India", 20, "Jun 16"),
    ("Rakesh", "Market Research, Cashback (co-owned), Survey Platforms",
     602 + 71 + 9, "Apr 23 to Jun 23"),
]

CHANGELOG = [
    ("Date", "Change", "Owner"),
    ("2026-04-29", "v4: Coverage Map now lists every one of the 30 Dream List labels "
     "exactly as in source (typos preserved). Nothing collapsed silently.", "Naitik"),
    ("2026-04-29", "v4: Online Gaming & Casino Loyalty gets its own slot "
     "(11 cos covering 10 sub-labels + Gaming gift card re-sellers).", "Naitik"),
    ("2026-04-29", "v4: Survey Platforms gets its own slot (Mon Jun 23, Rakesh).", "Naitik"),
    ("2026-04-29", "v4: BAAS (ITIO, 1 co) explicitly listed as folded into "
     "Remittance + Fintech BFSI Global combo.", "Naitik"),
    ("2026-04-29", "v3: marked 3 LIVE — Passive Deal, Customer Loyalty (Global API), Cashback.", "Naitik"),
    ("2026-04-29", "v3: ABM Touchpoints sheet with Smartlead numbers; HeyReach + HubSpot TBD.", "Naitik"),
    ("2026-04-29", "v2: 6 launch slots added covering 7 missing use cases.", "Naitik"),
    ("2026-04-29", "OPEN: 247 unassigned rows in Dream List need triage before launch.", "Naitik"),
    ("2026-04-29", "OPEN: HeyReach + HubSpot touchpoint counts to be filled manually for Live campaigns.", "Naitik"),
]


thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def title_cell(ws, cell_range, text):
    first = cell_range.split(":")[0]
    ws[first] = text
    ws[first].font = Font(bold=True, size=14, color="FFFFFF")
    ws[first].fill = PatternFill("solid", fgColor="1F4E78")
    ws[first].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(cell_range)


def subtitle_cell(ws, cell_range, text):
    first = cell_range.split(":")[0]
    ws[first] = text
    ws[first].font = Font(italic=True, color="555555")
    ws.merge_cells(cell_range)


def header_cell(cell, text, fg="305496"):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def main():
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Rollout Schedule
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Rollout Schedule"
    title_cell(ws, "A1:I1", "API HUNTING ABM — CAMPAIGN ROLLOUT v4")
    subtitle_cell(ws, "A2:I2",
        "21 launch slots covering all 30 Dream List use case labels. "
        "3 currently LIVE (see ABM Touchpoints sheet). 2 sequences live per week (Mon + Thu).")
    ws.row_dimensions[1].height = 24

    headers = ["Week", "Launch Day", "Campaign", "Priority",
               "# Companies", "Sales POC", "Status", "Source", "Notes"]
    for col, h in enumerate(headers, start=1):
        header_cell(ws.cell(row=4, column=col), h)

    status_fill = {
        "Live": PatternFill("solid", fgColor="C6EFCE"),
        "In Progress": PatternFill("solid", fgColor="FFEB9C"),
        "Planned": PatternFill("solid", fgColor="DDEBF7"),
        "New": PatternFill("solid", fgColor="F8CBAD"),
    }
    priority_fill = {
        "P0": PatternFill("solid", fgColor="C00000"),
        "P1": PatternFill("solid", fgColor="ED7D31"),
        "P2": PatternFill("solid", fgColor="70AD47"),
    }

    for ridx, row in enumerate(ROWS, start=5):
        for cidx, val in enumerate(row, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        pcell = ws.cell(row=ridx, column=4)
        pcell.fill = priority_fill.get(row[3], PatternFill())
        pcell.font = Font(bold=True, color="FFFFFF")
        pcell.alignment = Alignment(horizontal="center", vertical="center")
        scell = ws.cell(row=ridx, column=7)
        scell.fill = status_fill.get(row[6], PatternFill())
        scell.font = Font(bold=True)
        scell.alignment = Alignment(horizontal="center", vertical="center")
        if row[6] == "Live":
            for c in range(1, 4):
                ws.cell(row=ridx, column=c).fill = PatternFill("solid", fgColor="E2EFDA")
        elif row[6] == "New":
            for c in range(1, 4):
                ws.cell(row=ridx, column=c).fill = PatternFill("solid", fgColor="FFF2CC")

    widths = [10, 14, 50, 10, 13, 22, 14, 10, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    total_row = len(ROWS) + 5
    ws.cell(row=total_row, column=3, value="TOTAL (P0 + P1 + P2)").font = Font(bold=True)
    numeric = sum(r[4] for r in ROWS if isinstance(r[4], int))
    ws.cell(row=total_row, column=5, value=numeric).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value="P1 Fintech/BFSI count pending from Gaurav").font = Font(italic=True)
    for c in range(1, 10):
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")

    # ------------------------------------------------------------------
    # Sheet 2: ABM Touchpoints
    # ------------------------------------------------------------------
    ws_tp = wb.create_sheet("ABM Touchpoints")
    title_cell(ws_tp, "A1:M1", "ABM TOUCHPOINTS — LIVE CAMPAIGNS")
    subtitle_cell(ws_tp, "A2:M2",
        "Smartlead numbers pulled live on 2026-04-29. "
        "HeyReach + HubSpot columns = manual pull (no MCP available). "
        "Open/Click are 0 by design — tracking disabled to protect deliverability.")
    ws_tp.row_dimensions[1].height = 22

    for ridx, row in enumerate(TOUCHPOINTS, start=4):
        for cidx, val in enumerate(row, start=1):
            cell = ws_tp.cell(row=ridx, column=cidx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
            if ridx == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
            else:
                if cidx == 3:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(bold=True)
                if cidx in (10, 11, 12) and isinstance(val, str) and "TBD" in val:
                    cell.fill = PatternFill("solid", fgColor="FFE699")
                    cell.font = Font(italic=True)

    tp_widths = [32, 18, 22, 18, 18, 9, 9, 9, 9, 16, 16, 14, 60]
    for i, w in enumerate(tp_widths, start=1):
        ws_tp.column_dimensions[get_column_letter(i)].width = w
    ws_tp.freeze_panes = "B5"

    # ------------------------------------------------------------------
    # Sheet 3: Coverage Map (every Dream List label visible)
    # ------------------------------------------------------------------
    ws_cov = wb.create_sheet("Coverage Map")
    title_cell(ws_cov, "A1:F1", "COVERAGE MAP — ALL 30 DREAM LIST USE CASES")
    subtitle_cell(ws_cov, "A2:F2",
        "Every label from the Dream List preserved verbatim (including source typos). "
        "Each label is mapped to a launch slot. Single-entry edge cases are visible "
        "and explicitly labelled 'Folded in' or 'NEW slot — combined'.")
    ws_cov.row_dimensions[1].height = 22

    for ridx, row in enumerate(COVERAGE, start=4):
        for cidx, val in enumerate(row, start=1):
            cell = ws_cov.cell(row=ridx, column=cidx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if ridx == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if ridx > 4:
            status = row[5] if len(row) > 5 else ""
            if isinstance(status, str):
                if "TRIAGE" in status or "Open" in status:
                    fill = PatternFill("solid", fgColor="F4CCCC")
                elif "LIVE" in status:
                    fill = PatternFill("solid", fgColor="C6EFCE")
                elif "NEW" in status:
                    fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    fill = None
                if fill:
                    for c in range(1, 7):
                        ws_cov.cell(row=ridx, column=c).fill = fill

    cov_widths = [60, 8, 50, 14, 18, 36]
    for i, w in enumerate(cov_widths, start=1):
        ws_cov.column_dimensions[get_column_letter(i)].width = w
    ws_cov.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Sheet 4: POC Summary
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("POC Summary")
    title_cell(ws2, "A1:D1", "OWNERSHIP BY SALES POC")
    ws2.row_dimensions[1].height = 22
    for ridx, row in enumerate(POC_NOTES, start=3):
        for cidx, val in enumerate(row, start=1):
            cell = ws2.cell(row=ridx, column=cidx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if ridx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate([18, 80, 18, 28], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    # ------------------------------------------------------------------
    # Sheet 5: Change Log
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Change Log")
    title_cell(ws3, "A1:C1", "CHANGE LOG")
    ws3.row_dimensions[1].height = 22
    for ridx, row in enumerate(CHANGELOG, start=3):
        for cidx, val in enumerate(row, start=1):
            cell = ws3.cell(row=ridx, column=cidx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if ridx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([14, 100, 14], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A4"

    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Schedule rows: {len(ROWS)}  Total companies (P0+P1+P2): {numeric}")
    print(f"Live campaigns: 3 (Passive Deal, Customer Loyalty Global API, Cashback)")
    print(f"Dream List labels covered: {len(COVERAGE) - 2}  (excluding header and 247 unassigned row)")


if __name__ == "__main__":
    main()
